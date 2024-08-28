#!/usr/bin/python
from openpyxl import Workbook
import datetime
from openpyxl.styles import PatternFill
from timetable import profileDataForty, profileDataFifteen
wb = Workbook()
ws, count, profCount = wb.active, 2, 0
timeNow = datetime.datetime.now()
timeNow = timeNow.replace(hour=0, minute=0, second=0, microsecond=0)
timeList, profileToTags, profileNodeData = [], {}, {}
dataDesc=open("forty.txt","r+")
for slot in range(2, 290):
	tempTime = str(timeNow).split(' ')[1].split(":")[0]+":"+str(timeNow).split(' ')[1].split(":")[1]
	timeList.append(tempTime)
	ws.cell(column=slot, row=1, value=tempTime)
	timeNow = timeNow + datetime.timedelta(minutes=5)
sleepList, timedList = [], []
for line in dataDesc.readlines():
    if line !="\n":
	if "\t" not in line:
		ws['A'+str(count)] = line.strip("\n")
		count+=1
		profCount+=1
		Profile = line.strip("\n").split(":")[0].strip(" ")
		ProfileTags = ['0'] * len(timeList)
		nodeData = []
	else:
		if "SCHEDULE_SLEEP" in line:
			slotDiff = int(line.split(" ")[1].strip("\n"))/(300)
			if slotDiff != 0:
			    for i in range(0,len(timeList),slotDiff):
				ProfileTags[i]=1
			profileToTags[Profile] = ProfileTags
			sleepList.append(Profile)
		if "SCHEDULED_TIMES_STRINGS" in line:
			slotTimes = line.split(" ")[1].strip("'").strip("[").strip("]\n").split(", ")
			for slots in slotTimes:
				ProfileTags[timeList.index(slots.strip("'").split(":")[0]+":"+slots.strip("'").split(":")[1])]=1
			profileToTags[Profile] = ProfileTags
			timedList.append(Profile)
		if "TOTAL_NODES" in line:
			nodeData.append(line)
		if "NUM_NODES" in line:
			nodeData.append(line)
		if "SUPPORTED_NODE_TYPES" in line:
			nodeData.append(line)
		nodeData=list(set(nodeData))
		profileNodeData[Profile] = nodeData
for rowSlot in range(2,291):
	profile = ws["A"+str(rowSlot)].value.strip(" ")
	if profile in sleepList:
		for columnSlot in range(2, 290):
		    try:
		         ws.cell(column=columnSlot, row=rowSlot, value=1).fill = PatternFill(fgColor="00FF0000", fill_type = "solid")
		    except:	
			pass
        #ws["B"+str(rowSlot)]=profileNodeData[profile]
	for columnSlot in range(2, 290):
	    try:
		if profileToTags[profile][columnSlot-2] == 1:
 		    ws.cell(column=columnSlot, row=rowSlot, value=profileToTags[profile][columnSlot-2]).fill = PatternFill(fgColor="00FF0000", fill_type = "solid")
	    except:
		pass
runTimeDict={}
rowNum = ws.max_row
for columnSlot in range(2, ws.max_column):
	count = 0
	for rowSlot in range(2,ws.max_row):
	    if ws.cell(row=rowSlot, column=columnSlot).value != None:
		count += ws.cell(row=rowSlot, column=columnSlot).value
	runTimeDict[ws.cell(row=1, column=columnSlot).value]=count
	ws.cell(column=columnSlot, row=rowNum+1, value=count)
wb.save("sample.xlsx")
runTimes, listToDictMap, count = runTimeDict.values(),{},0
runTimes.sort(reverse=True)
for runs in runTimes[0:4]:
    for item in runTimeDict:
	if runTimeDict[item] == runs:
	    profileListAtInstant=[]
  	    for rowSlot in range(2,ws.max_row+2):
		if ws.cell(row=rowSlot,column=timeList.index(item)+2).value == 1:
		    profileListAtInstant.append(str(ws.cell(row=rowSlot, column=1).value.split(" :")[0]))
    	    listToDictMap[count]=profileListAtInstant
    	    count+=1
for itemList in listToDictMap:
	print "#######################################"
	for profile in listToDictMap[itemList]:
		profile=profile.strip(" ")
		if profile in profileDataFifteen.keys():
			if "NUM_NODES" in profileDataFifteen[profile] and "NUM_NODES=NA" not in profileDataFifteen[profile]:
				print profile
				for item in profileDataFifteen[profile].split("\n"):
				    if "NUM_NODES" in item:
					if "NUM_NODES=NA" not in item:
 					    print item
				    if "TOTAL_NODES" in item:
					print item
				    if "SUPPORTED_NODE_TYPES" in item:
					print item
			if "TOTAL_NODES" in profileDataFifteen[profile]:
                                print profile
                                for item in profileDataFifteen[profile].split("\n"):
                                    if "NUM_NODES" in item:
                                        if "NUM_NODES=NA" not in item:
                                            print item
                                    if "TOTAL_NODES" in item:
                                        print item
                                    if "SUPPORTED_NODE_TYPES" in item:
                                        print item
		elif profile in profileDataForty.keys():
                       if "NUM_NODES" in profileDataForty[profile] and "NUM_NODES=NA" not in profileDataForty[profile]:
                                print profile
                                for item in profileDataForty[profile].split("\n"):
                                    if "NUM_NODES" in item:
                                        if "NUM_NODES=NA" not in item:
                                            print item
                                    if "TOTAL_NODES" in item:
                                        print item
                                    if "SUPPORTED_NODE_TYPES" in item:
                                        print item
                       if "TOTAL_NODES" in profileDataForty[profile]:
                                print profile
                                for item in profileDataForty[profile].split("\n"):
                                    if "NUM_NODES" in item:
                                        if "NUM_NODES=NA" not in item:
                                            print item
                                    if "TOTAL_NODES" in item:
                                        print item
                                    if "SUPPORTED_NODE_TYPES" in item:
                                        print item
